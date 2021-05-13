from process_functions import *
import openpyxl

class MainProcess(Process):

    def __init__(self, file_name=""):
        super().__init__()
        self.index = 0
        self.dis_array = []
        self.ind_array = []
        self.video_address = file_name
        self.final_width = 620
        self.final_height = 180
        self.distance_x = 0
        self.number_of_frame = 4


    def get_folder_address(self):
        path_video = ""
        is_path = False
        path = self.video_address.split('.')
        if path[-1] == "mp4":
            path = path[-2].split('/')
            for i in range(len(path) - 1):
                if path[i] == 'Inputs':
                    is_path = True
                    continue
                if is_path:
                    path_video += path[i] + '/'
            # path_excel = path_video
            path_video += path[-1]
        # print("Path_video_in_project:", path_video)

        # elif path[1] == "jpg":
        #     is_path = False
        #     path = path[0].split('/')
        #     for i in range(len(path) - 3):
        #         if path[i] == 'Outputs':
        #             is_path = True
        #             continue
        #         if is_path:
        #             path_video += path[i] + '/'
        #     path_video += path[-3]
        #     # print(path_video)

        image_subfolder = []
        frame_folder = 'Outputs/' + path_video + '/frameFolder'
        image_folder = 'Outputs/' + path_video + '/imageFolder'
        [image_subfolder.append(image_folder + '/X{}_Folder'.format(i+1)) for i in range(self.number_of_frame)]
        excel_folder = 'Outputs/' + path_video + '/excelFolder'
        return frame_folder, image_folder, excel_folder, image_subfolder


    def check_folder(self):
        frame_folder, image_folder, excel_folder, image_subfolder = self.get_folder_address()

        if not os.path.exists(frame_folder):
            os.makedirs(frame_folder)
        if not os.path.exists(excel_folder):
            os.makedirs(excel_folder)
        if os.path.exists(image_folder):
            shutil.rmtree(image_folder)
        for subfolder in image_subfolder:
            os.makedirs(subfolder)

    def get_save_address(self):
        _, _, excel_folder, _ = self.get_folder_address()
        save_address = excel_folder + '/time-distance.xlsx'
        return save_address

    def get_index(self):
        frame_folder, _, _, _ = self.get_folder_address()
        if self.index == 0:
            path, dirs, files = next(os.walk(frame_folder))
            return len(files)
        else:
            return self.index

    # def get_image_address(self, ind_image):
    #     _, img_folder, _, _ = self.get_folder_address()
    #     img_address = []
    #     for filename in glob.glob(img_folder + '/*.jpg'):
    #         img_address.append(filename)
    #         # print(filename)
    #     return img_address[ind_image]

    # def get_image(self, ind_image):
    #     _, image_folder, _ = self.get_folder_address()
    #     image_address = image_folder

    def get_frame(self, ind_frame):
        frame_folder, _, _, _ = self.get_folder_address()
        frame_address = frame_folder + '/Frame' + str('{0:04}'.format(ind_frame) + '.jpg')
        return cv2.imread(frame_address)

    def sort_option(self, option):
        return option[0] #Sort bases on position of frame

    def print_progress_bar(self, iterations, suffix=''):
        # print("Frame: {}".format(iterations), end =" ")
        total = self.get_index()
        percent = ("{0:." + str(1) + "f}").format(100 * ((iterations + 1) / float(total)))
        filled_length = int(100 * iterations // total)
        bar = '#' * filled_length + "Frame: {}".format(iterations) +'-' * (100 - filled_length)
        print('Process: |{}| {}%'.format(bar, percent))
        # print('\r Process: |%s| %s%% \r\n' % (bar, percent))
        # Print new line on complete
        if iterations == total:
            print()

    def process_image(self, ind_image):
        self.print_progress_bar(ind_image)

        # STEP 1: Load image
        original_image = self.get_frame(ind_image)
        original_image = cv2.rotate(original_image, cv2.ROTATE_90_CLOCKWISE)
        # cv2.imshow("Original image", cv2.resize(src=original_image, dsize=(500, 200)))

        # STEP 2: Detect WHITE frame
        set_of_white_frame = super().detect_white_frame(original_image)
        if np.shape(set_of_white_frame)[0] != self.number_of_frame:
            print("[ERROR] in STEP 2")
            self.error_list["step2"] = True
            return
        set_of_white_frame.sort(key=self.sort_option)
        i = 0
        image_distance = []
        for position, white_frame in set_of_white_frame:
            white_frame = cv2.resize(white_frame, (self.final_width, self.final_height))
            # cv2.imshow("Detected white frame", white_frame)
            # cv2.imwrite("detect.jpg", white_frame)

            # STEP 3: Determine the coordinate of the Grid
            threshold_grid = super().determine_threshold_for_grid(white_frame)
            super().set_threshold(threshold_grid)

            line1, line2, ver_coor, translation = super().detect_grid_coodinate(white_frame)

            # STEP 4: Find center point
            cX, cY = super().find_center_point(white_frame)
            if cX == -1:
                print("[ERROR] in STEP 4")
                self.error_list["step4"] = True
                return

            # STEP 5: Calculate the real coordinate of the laser pointer
            self.distance_x = super().calculate_real_coordinate_of_laser_pointer(cX, cY, ver_coor)
            print("Distance of X{}: {}".format(i+1, self.distance_x))

            # STEP 6: Draw and Save image
            final_image = super().shift_image(white_frame, translation)
            font = cv2.FONT_HERSHEY_COMPLEX
            for j in range(15):
                cv2.line(final_image, (line1[j][0] + translation, line1[j][1]), (line2[j][0], line2[j][1]),
                         (0, 0, 0), 1)

            cv2.circle(final_image, (cX, cY), 5, (0, 0, 255), 1, cv2.LINE_AA)
            cv2.line(final_image, (cX, cY), (ver_coor[0], cY), (0, 0, 255), 1)
            cv2.putText(final_image, str(self.distance_x) + 'cm', (int(cX / 2), cY - 10), font, 0.5, (255, 0, 0))
            final_img_size = np.shape(final_image)

            final_image = cv2.resize(src=final_image[:, abs(translation):final_img_size[1] - abs(translation), :],
                                     dsize=(self.final_width, self.final_height))

            _, _, _, image_subfolder = self.get_folder_address()
            image_address = image_subfolder[i] + '/Image' + str(
                '{0:04}-{distance}'.format(ind_image, distance=self.distance_x)) + '.jpg'
            # print(image_address)
            i += 1
            cv2.imwrite(image_address, final_image)
            image_distance.append(self.distance_x)
        self.dis_array.append(image_distance)
        self.ind_array.append(ind_image)


    def save_excel_file(self):
        first_row = ['Frame']
        if not os.path.exists(self.get_save_address()):
            workbook = xlsxwriter.Workbook(self.get_save_address())
            worksheet = workbook.add_worksheet('result')
            [first_row.append('X{}'.format(i+1)) for i in range(self.number_of_frame)]
            worksheet.write_row(0, 0, first_row)
            workbook.close()

        wb = openpyxl.load_workbook(self.get_save_address())
        ws = wb['result']
        n = self.number_of_frame
        for i in range(np.shape(self.ind_array)[0]):
            ws.cell(self.ind_array[i] + 1, 1).value = self.ind_array[i] #Frame
            for j in range(n):
                ws.cell(self.ind_array[i] + 1 + 1, j + 2).value = self.dis_array[i][j]
        wb.save(self.get_save_address())
        wb.close()


