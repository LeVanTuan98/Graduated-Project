# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'laserProcessGUI.ui'
#
# Created by: PyQt5 UI code generator 5.13.0
#
# WARNING! All changes made in this file will be lost!


import pyqtgraph

import xlrd
import openpyxl

from main_process import *
from main_frame import *
from custome_dialog import *
from my_thread import *


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1081, 625)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        MainWindow.setWindowIcon(QtGui.QIcon('Resourses/Icons/balance.svg'))

        # -------------------------Infomation Box-----------------------#
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(10, 10, 261, 371))
        self.groupBox.setObjectName("groupBox")
        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setGeometry(QtCore.QRect(20, 30, 91, 16))
        self.label.setObjectName("label")
        self.id_txt = QtWidgets.QLineEdit(self.groupBox)
        self.id_txt.setGeometry(QtCore.QRect(130, 30, 113, 20))
        self.id_txt.setObjectName("id_txt")
        self.label_2 = QtWidgets.QLabel(self.groupBox)
        self.label_2.setGeometry(QtCore.QRect(20, 60, 91, 16))
        self.label_2.setObjectName("label_2")
        self.name_txt = QtWidgets.QLineEdit(self.groupBox)
        self.name_txt.setGeometry(QtCore.QRect(130, 60, 113, 20))
        self.name_txt.setObjectName("name_txt")
        self.label_3 = QtWidgets.QLabel(self.groupBox)
        self.label_3.setGeometry(QtCore.QRect(20, 90, 91, 16))
        self.label_3.setObjectName("label_3")
        self.age_txt = QtWidgets.QLineEdit(self.groupBox)
        self.age_txt.setGeometry(QtCore.QRect(130, 90, 113, 20))
        self.age_txt.setObjectName("age_txt")
        self.label_4 = QtWidgets.QLabel(self.groupBox)
        self.label_4.setGeometry(QtCore.QRect(20, 120, 91, 21))
        self.label_4.setObjectName("label_4")
        self.sex_opt = QtWidgets.QComboBox(self.groupBox)
        self.sex_opt.setGeometry(QtCore.QRect(130, 120, 111, 22))
        self.sex_opt.setObjectName("sex_opt")
        self.sex_opt.addItem("")
        self.sex_opt.addItem("")
        self.label_5 = QtWidgets.QLabel(self.groupBox)
        self.label_5.setGeometry(QtCore.QRect(20, 150, 91, 21))
        self.label_5.setObjectName("label_5")
        self.height_txt = QtWidgets.QLineEdit(self.groupBox)
        self.height_txt.setGeometry(QtCore.QRect(130, 150, 113, 20))
        self.height_txt.setObjectName("height_txt")
        self.label_6 = QtWidgets.QLabel(self.groupBox)
        self.label_6.setGeometry(QtCore.QRect(20, 180, 91, 21))
        self.label_6.setObjectName("label_6")
        self.weight_txt = QtWidgets.QLineEdit(self.groupBox)
        self.weight_txt.setGeometry(QtCore.QRect(130, 180, 113, 20))
        self.weight_txt.setObjectName("weight_txt")
        self.add_txt = QtWidgets.QLineEdit(self.groupBox)
        self.add_txt.setGeometry(QtCore.QRect(20, 230, 221, 41))
        self.add_txt.setObjectName("add_txt")
        self.label_7 = QtWidgets.QLabel(self.groupBox)
        self.label_7.setGeometry(QtCore.QRect(20, 210, 91, 21))
        self.label_7.setObjectName("label_7")
        self.note_txt = QtWidgets.QLineEdit(self.groupBox)
        self.note_txt.setGeometry(QtCore.QRect(20, 300, 221, 51))
        self.note_txt.setObjectName("note_txt")
        self.label_8 = QtWidgets.QLabel(self.groupBox)
        self.label_8.setGeometry(QtCore.QRect(20, 270, 91, 21))
        self.label_8.setObjectName("label_8")

        # ---------------------------Frame Box------------------------#
        self.groupBox_2 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_2.setGeometry(QtCore.QRect(290, 10, 761, 431))
        self.groupBox_2.setObjectName("groupBox_2")
        self.scrollBar = QtWidgets.QScrollBar(self.groupBox_2)
        self.scrollBar.setGeometry(QtCore.QRect(10, 29, 20, 371))
        self.scrollBar.setOrientation(QtCore.Qt.Vertical)
        self.scrollBar.setObjectName("scrollBar")
        self.scrollBar.setMaximum(0)
        self.main_frame = MainFrame(self.groupBox_2)
        self.main_frame.setGeometry(QtCore.QRect(216, 20, 531, 381))
        self.main_frame.setText("")
        self.main_frame.setObjectName("main_frame")
        self.sub_frame1 = SubFrame(self.groupBox_2)
        self.sub_frame1.setGeometry(QtCore.QRect(50, 20, 141, 71))
        self.sub_frame1.setText("")
        self.sub_frame1.setObjectName("sub_frame1")
        self.sub_frame2 = SubFrame(self.groupBox_2)
        self.sub_frame2.setGeometry(QtCore.QRect(50, 110, 141, 71))
        self.sub_frame2.setText("")
        self.sub_frame2.setObjectName("sub_frame2")
        self.sub_frame3 = SubFrame(self.groupBox_2)
        self.sub_frame3.setGeometry(QtCore.QRect(50, 200, 141, 81))
        self.sub_frame3.setText("")
        self.sub_frame3.setObjectName("sub_frame3")
        self.sub_frame4 = SubFrame(self.groupBox_2)
        self.sub_frame4.setGeometry(QtCore.QRect(50, 300, 141, 81))
        self.sub_frame4.setText("")
        self.sub_frame4.setObjectName("sub_frame4")

        # ---------------------------Control Panel----------------------#
        self.groupBox_3 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_3.setGeometry(QtCore.QRect(10, 390, 261, 161))
        self.groupBox_3.setObjectName("groupBox_3")
        self.calib_btn = QtWidgets.QPushButton(self.groupBox_3)
        self.calib_btn.setGeometry(QtCore.QRect(20, 20, 61, 23))
        self.calib_btn.setObjectName("calib_btn")
        self.run_btn = QtWidgets.QPushButton(self.groupBox_3)
        self.run_btn.setGeometry(QtCore.QRect(20, 60, 61, 23))
        self.run_btn.setObjectName("run_btn")
        # self.chart_rbtn = QtWidgets.QRadioButton(self.groupBox_3)
        # self.chart_rbtn.setGeometry(QtCore.QRect(110, 20, 82, 17))
        # self.chart_rbtn.setObjectName("chart_rbtn")
        self.label_9 = QtWidgets.QLabel(self.groupBox_3)
        self.label_9.setGeometry(QtCore.QRect(110, 20, 65, 17))
        self.label_9.setObjectName("label_9")
        self.direc_opt = QtWidgets.QComboBox(self.groupBox_3)
        self.direc_opt.setGeometry(QtCore.QRect(180, 20, 50, 17))
        self.direc_opt.setObjectName("direc_opt")
        self.direc_opt.addItem("")
        self.direc_opt.addItem("")

        self.progressBar = QtWidgets.QProgressBar(self.groupBox_3)
        self.progressBar.setGeometry(QtCore.QRect(100, 63, 151, 20))
        self.progressBar.setProperty("value", 24)
        self.progressBar.setObjectName("progressBar")
        self.progressBar.setMaximum(100)
        self.progressBar.setStyleSheet("QProgressBar {border: 1px solid grey; border-radius:5px; padding:1px}")
        self.progressBar.setValue(0)

        # --------------------------Console Box-----------------------#
        self.groupBox_4 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_4.setGeometry(QtCore.QRect(290, 439, 761, 111))
        self.groupBox_4.setObjectName("groupBox_4")
        self.console_list = QtWidgets.QListWidget(self.groupBox_4)
        self.console_list.setGeometry(QtCore.QRect(20, 20, 731, 81))
        self.console_list.setObjectName("console_list")

        # ----------------------------Menu----------------------------#
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1081, 21))
        self.menubar.setObjectName("menubar")
        self.menu_File = QtWidgets.QMenu(self.menubar)
        self.menu_File.setObjectName("menu_File")
        self.menu_Open_Options = QtWidgets.QMenu(self.menu_File)
        self.menu_Open_Options.setObjectName("menu_Open_Options")
        self.menu_Save_Options = QtWidgets.QMenu(self.menu_File)
        self.menu_Save_Options.setObjectName("menu_Save_Options")
        self.menu_Edit = QtWidgets.QMenu(self.menubar)
        self.menu_Edit.setObjectName("menu_Edit")
        self.menu_Help = QtWidgets.QMenu(self.menubar)
        self.menu_Help.setObjectName("menu_Help")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.toolBar = QtWidgets.QToolBar(MainWindow)
        self.toolBar.setObjectName("toolBar")
        MainWindow.addToolBar(QtCore.Qt.TopToolBarArea, self.toolBar)
        self.action_New = QtWidgets.QAction(MainWindow)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("Resourses/Icons/new.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_New.setIcon(icon)
        self.action_New.setObjectName("action_New")
        self.action_Open = QtWidgets.QAction(MainWindow)
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("Resourses/Icons/open.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_Open.setIcon(icon1)
        self.action_Open.setObjectName("action_Open")
        self.action_Save = QtWidgets.QAction(MainWindow)
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap("Resourses/Icons/save.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_Save.setIcon(icon2)
        self.action_Save.setObjectName("action_Save")
        self.action_Exit = QtWidgets.QAction(MainWindow)
        icon3 = QtGui.QIcon()
        icon3.addPixmap(QtGui.QPixmap("Resourses/Icons/exit.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_Exit.setIcon(icon3)
        self.action_Exit.setObjectName("action_Exit")
        self.action_Open_image_folder = QtWidgets.QAction(MainWindow)
        self.action_Open_image_folder.setObjectName("action_Open_image_folder")
        # self.action_Open_chart = QtWidgets.QAction(MainWindow)
        # self.action_Open_chart.setObjectName("action_Open_chart")
        self.menu_Open_chart = QtWidgets.QMenu(MainWindow)
        self.menu_Open_chart.setObjectName("menu_Open_chart")

        self.action_Copy = QtWidgets.QAction(MainWindow)
        icon4 = QtGui.QIcon()
        icon4.addPixmap(QtGui.QPixmap("Resourses/Icons/copy.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_Copy.setIcon(icon4)
        self.action_Copy.setObjectName("action_Copy")
        self.action_Paste = QtWidgets.QAction(MainWindow)
        icon5 = QtGui.QIcon()
        icon5.addPixmap(QtGui.QPixmap("Resourses/Icons/paste.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_Paste.setIcon(icon5)
        self.action_Paste.setObjectName("action_Paste")
        self.action_Cut = QtWidgets.QAction(MainWindow)
        icon6 = QtGui.QIcon()
        icon6.addPixmap(QtGui.QPixmap("Resourses/Icons/cut.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_Cut.setIcon(icon6)
        self.action_Cut.setObjectName("action_Cut")
        self.action_File_and_Replace = QtWidgets.QAction(MainWindow)
        self.action_File_and_Replace.setObjectName("action_File_and_Replace")
        self.action_About = QtWidgets.QAction(MainWindow)
        icon7 = QtGui.QIcon()
        icon7.addPixmap(QtGui.QPixmap("Resourses/Icons/about.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_About.setIcon(icon7)
        self.action_About.setObjectName("action_About")
        self.action_Save_excel_file = QtWidgets.QAction(MainWindow)
        self.action_Save_excel_file.setObjectName("action_Save_excel_file")
        self.action_Save_chart = QtWidgets.QAction(MainWindow)
        self.action_Save_chart.setObjectName("action_Save_chart")
        self.action_Zoom_In = QtWidgets.QAction(MainWindow)
        icon8 = QtGui.QIcon()
        icon8.addPixmap(QtGui.QPixmap("Resourses/Icons/zoomin.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_Zoom_In.setIcon(icon8)
        self.action_Zoom_In.setObjectName("action_Zoom_In")
        self.action_Zoom_Out = QtWidgets.QAction(MainWindow)
        icon9 = QtGui.QIcon()
        icon9.addPixmap(QtGui.QPixmap("Resourses/Icons/zoomout.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_Zoom_Out.setIcon(icon9)
        self.action_Zoom_Out.setObjectName("action_Zoom_Out")
        self.menu_Open_Options.addAction(self.action_Open_image_folder)
        # self.menu_Open_Options.addAction(self.action_Open_chart)
        self.menu_Open_Options.addAction(self.menu_Open_chart.menuAction())
        self.menu_Save_Options.addAction(self.action_Save_excel_file)
        self.menu_Save_Options.addAction(self.action_Save_chart)
        self.menu_File.addAction(self.action_New)
        self.menu_File.addSeparator()
        self.menu_File.addAction(self.action_Open)
        self.menu_File.addAction(self.menu_Open_Options.menuAction())

        self.action_Open_X = QtWidgets.QAction(MainWindow)
        self.action_Open_X.setObjectName("action_Open_X")
        self.action_Open_Y = QtWidgets.QAction(MainWindow)
        self.action_Open_Y.setObjectName("action_Open_Y")
        self.action_Open_XY = QtWidgets.QAction(MainWindow)
        self.action_Open_XY.setObjectName("action_Open_XY")
        self.menu_Open_chart.addAction(self.action_Open_X)
        self.menu_Open_chart.addAction(self.action_Open_Y)
        self.menu_Open_chart.addAction(self.action_Open_XY)

        self.menu_File.addSeparator()
        self.menu_File.addAction(self.action_Save)
        self.menu_File.addAction(self.menu_Save_Options.menuAction())
        self.menu_File.addSeparator()
        self.menu_File.addAction(self.action_Exit)
        self.menu_Edit.addAction(self.action_Copy)
        self.menu_Edit.addAction(self.action_Paste)
        self.menu_Edit.addAction(self.action_Cut)
        self.menu_Edit.addSeparator()
        self.menu_Edit.addAction(self.action_Zoom_In)
        self.menu_Edit.addAction(self.action_Zoom_Out)
        self.menu_Edit.addSeparator()
        self.menu_Edit.addAction(self.action_File_and_Replace)
        self.menu_Help.addAction(self.action_About)
        self.menubar.addAction(self.menu_File.menuAction())
        self.menubar.addAction(self.menu_Edit.menuAction())
        self.menubar.addAction(self.menu_Help.menuAction())
        self.toolBar.addAction(self.action_New)
        self.toolBar.addAction(self.action_Open)
        self.toolBar.addAction(self.action_Save)
        self.toolBar.addAction(self.action_Exit)
        self.toolBar.addSeparator()
        self.toolBar.addAction(self.action_Zoom_In)
        self.toolBar.addAction(self.action_Zoom_Out)
        self.toolBar.addSeparator()
        self.toolBar.addAction(self.action_About)

        # --------------------------------Chart---------------------------------#
        self.graphicsView = pyqtgraph.PlotWidget(self.groupBox_2)
        self.graphicsView.setGeometry(QtCore.QRect(216, 20, 531, 381))
        # brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        # brush.setStyle(QtCore.Qt.NoBrush)
        # self.graphicsView.setBackgroundBrush(brush)
        color = self.graphicsView.palette().color(QtGui.QPalette.Window)
        self.graphicsView.setBackgroundBrush(color)
        self.graphicsView.setObjectName("graphicsView")
        self.graphicsView.setEnabled(True)
        self.graphicsView.setStyleSheet("background:white")
        # self.graphicsView.setTitle("ĐỒ THỊ DAO ĐỘNG", color="b", size="30pt")
        # self.graphicsView.setLabel('left', 'Khoang cach', units='cm')
        # self.graphicsView.setLabel('bottom', 'Thoi gian', units='s')
        self.graphicsView.setTitle("<span style=\"color:blue;font-size:18pt\">ĐỒ THỊ DAO ĐỘNG</span>")
        styles = {'color': 'b', 'font-size': '15px', 'padding': '0'}
        self.graphicsView.setLabel('left', 'Khoảng cách (cm)', **styles)
        self.graphicsView.setLabel('bottom', 'Thời gian (s)', **styles)
        self.graphicsView.showGrid(x=True, y=True)
        # self.graphicsView.setXRange(-10, 10, padding=0)
        self.graphicsView.setYRange(-5, 5, padding=0)
        self.graphicsView.setGeometry(QtCore.QRect(0, 0, 0, 0))

        # self.rel_path = ""
        self.abs_path = ""
        self.mode = 0 # mode: 0(video), 1(image)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "RLTB"))
        self.groupBox.setTitle(_translate("MainWindow", "Information"))
        self.label.setText(_translate("MainWindow", "Patient\'s ID:"))
        self.label_2.setText(_translate("MainWindow", "Patient\'s Name:"))
        self.label_3.setText(_translate("MainWindow", "Patient\'s Age:"))
        self.label_4.setText(_translate("MainWindow", "Patient\'s Sex:"))
        self.sex_opt.setItemText(0, _translate("MainWindow", "Male"))
        self.sex_opt.setItemText(1, _translate("MainWindow", "Female"))
        self.label_5.setText(_translate("MainWindow", "Patient\'s Height:"))
        self.label_6.setText(_translate("MainWindow", "Patient\'s Weight:"))
        self.label_7.setText(_translate("MainWindow", "Patient\'s Address:"))
        self.label_8.setText(_translate("MainWindow", "Notes:"))
        self.groupBox_2.setTitle(_translate("MainWindow", "Main Window"))
        self.groupBox_3.setTitle(_translate("MainWindow", "Control Panel"))
        self.calib_btn.setText(_translate("MainWindow", "Calibrate"))
        self.run_btn.setText(_translate("MainWindow", "Run..."))
        # self.chart_rbtn.setText(_translate("MainWindow", "Show chart"))
        self.label_9.setText(_translate("MainWindow", "Direction: "))
        self.direc_opt.setItemText(0, _translate("MainWindow", "X"))
        self.direc_opt.setItemText(1, _translate("MainWindow", "Y"))
        self.groupBox_4.setTitle(_translate("MainWindow", "Console Log"))
        self.menu_File.setTitle(_translate("MainWindow", "&File"))
        self.menu_Open_Options.setTitle(_translate("MainWindow", "&Open Options"))
        self.menu_Save_Options.setTitle(_translate("MainWindow", "&Save Options"))
        self.menu_Edit.setTitle(_translate("MainWindow", "&Edit"))
        self.menu_Help.setTitle(_translate("MainWindow", "&Help"))
        self.toolBar.setWindowTitle(_translate("MainWindow", "toolBar"))
        self.action_New.setText(_translate("MainWindow", "&New..."))
        self.action_New.setToolTip(_translate("MainWindow", "Process the new video file"))
        self.action_New.setShortcut(_translate("MainWindow", "Ctrl+N"))
        self.action_Open.setText(_translate("MainWindow", "&Open..."))
        self.action_Open.setToolTip(_translate("MainWindow", "Open the processed file"))
        self.action_Open.setShortcut(_translate("MainWindow", "Ctrl+O"))
        self.action_Save.setText(_translate("MainWindow", "&Save"))
        self.action_Save.setToolTip(_translate("MainWindow", "Save result"))
        self.action_Save.setShortcut(_translate("MainWindow", "Ctrl+S"))
        self.action_Exit.setText(_translate("MainWindow", "&Exit"))
        self.action_Exit.setToolTip(_translate("MainWindow", "Exit the window"))
        self.action_Exit.setShortcut(_translate("MainWindow", "Ctrl+E"))

        self.action_Open_image_folder.setText(_translate("MainWindow", "&Open Image Folder"))
        self.menu_Open_chart.setTitle(_translate("MainWindow", "&Open Chart"))
        self.action_Open_X.setText(_translate("MainWindow", "X"))
        self.action_Open_Y.setText(_translate("MainWindow", "Y"))
        self.action_Open_XY.setText(_translate("MainWindow", "X-Y"))

        self.action_Copy.setText(_translate("MainWindow", "&Copy"))
        self.action_Copy.setShortcut(_translate("MainWindow", "Ctrl+C"))
        self.action_Paste.setText(_translate("MainWindow", "&Paste"))
        self.action_Paste.setShortcut(_translate("MainWindow", "Ctrl+V"))
        self.action_Cut.setText(_translate("MainWindow", "&Cut"))
        self.action_Cut.setShortcut(_translate("MainWindow", "Ctrl+X"))
        self.action_File_and_Replace.setText(_translate("MainWindow", "&File and Replace..."))
        self.action_About.setText(_translate("MainWindow", "&About..."))
        self.action_Save_excel_file.setText(_translate("MainWindow", "&Save excel file"))
        self.action_Save_chart.setText(_translate("MainWindow", "&Save chart"))
        self.action_Zoom_In.setText(_translate("MainWindow", "&Zoom In"))
        self.action_Zoom_In.setShortcut(_translate("MainWindow", "Ctrl++"))
        self.action_Zoom_Out.setText(_translate("MainWindow", "&Zoom Out"))
        self.action_Zoom_Out.setShortcut(_translate("MainWindow", "Ctrl+-"))

        #----------------------------Menu control-------------------------------------------#
        self.action_Save.triggered.connect(self.do_save_action)
        self.action_Save_excel_file.triggered.connect(self.do_save_action)

        self.action_New.triggered.connect(self.do_new_action)
        self.action_Exit.triggered.connect(sys.exit)
        self.action_About.triggered.connect(lambda: self.show_about_action())

        self.action_Open.triggered.connect(self.do_open_action)
        self.action_Open_image_folder.triggered.connect(self.do_open_image_action)
        self.action_Open_X.triggered.connect(self.do_open_chart_action_X)
        self.action_Open_Y.triggered.connect(self.do_open_chart_action_Y)
        self.action_Open_XY.triggered.connect(self.do_open_chart_action_XY)


        # ----------------------------Control panel action----------------------------------#
        self.run_btn.clicked.connect(self.run_btn_clicked)
        self.calib_btn.clicked.connect(self.calib_btn_clicked)
        self.scrollBar.valueChanged.connect(lambda: self.update_scrollbar_value())

        #-------------------------------Frame Box-------------------------------------------#
        self.sub_frame1.clicked.connect(self.show_subframe_to_mainframe)
        self.sub_frame2.clicked.connect(self.show_subframe_to_mainframe)
        self.sub_frame3.clicked.connect(self.show_subframe_to_mainframe)
        self.sub_frame4.clicked.connect(self.show_subframe_to_mainframe)


    def start_run_thread(self):
        self.run_thread = RunThread()
        self.run_thread.set_direction(self.direc_opt.currentIndex())
        self.run_thread.set_blue_HSV(self.main_frame.blue_HSV)
        self.run_thread.set_laser_HSV(self.main_frame.laser_HSV)
        self.run_thread.set_file_path(self.abs_path)
        self.run_thread.change_value_progress.connect(self.set_progress_value)
        self.run_thread.update_console_list.connect(self.update_console_log)
        self.run_thread.data_graph.connect(self.draw_graph)
        self.run_thread.total_image.connect(self.set_max_scrollbar)
        self.run_thread.run_progess_finished.connect(self.show_finish_noti)
        self.run_thread.start()

    def start_calib_thread(self):
        self.calib_thread = CalibThread()
        self.calib_thread.set_file_path(self.abs_path)
        self.calib_thread.change_value_progress.connect(self.set_progress_value)
        self.calib_thread.calib_progess_finished.connect(self.do_calib_button)
        self.calib_thread.start()

    def show_about_action(self):
        dlg = About()
        if dlg.exec_():
            pass

    def update_scrollbar_value(self):
        # getting current value
        value = self.scrollBar.value()
        # print(value)
        self.show_sub_frame(value)

    def set_progress_value(self, value):
        self.progressBar.setValue(value)

    def set_max_scrollbar(self, max_value):
        self.scrollBar.setMaximum(max_value - 4)

    def get_info_patient(self):
        data = [["Patient's ID ", self.id_txt.text()],
                ["Patient's Name", self.name_txt.text()],
                ["Patient's Age", self.age_txt.text()],
                ["Patient's Sex", self.sex_opt.currentText()],
                ["Patient's Height", self.height_txt.text()],
                ["Patient's Weight", self.weight_txt.text()],
                ["Patient's Address", self.add_txt.text()],
                ["Notes", self.note_txt.text()]]
        return data

    def do_new_action(self):
        self.abs_path, _ = QFileDialog.getOpenFileName(None, "Open Video File", 'Inputs/', "(*.mp4)")
        print("Path got video:", self.abs_path)
        self.mode = 0
        self.enable_graph()
        self.clear_frame_box()
        # self.disable_graph()
        self.set_infomation_box(["", "", "", "Male", "", "", "", ""])
        self.console_list.clear()

    def do_open_chart_action_X(self):
        self.do_open_chart_action(0)

    def do_open_chart_action_Y(self):
        self.do_open_chart_action(1)

    def do_open_chart_action_XY(self):
        self.do_open_chart_action(2)


    def do_open_action(self):
        self.abs_path = QFileDialog.getExistingDirectory(None, "Open folder", "Outputs/")
        print("Path got folder:", self.abs_path)
        self.console_list.clear()
        self.mode = 1
        self.set_max_scrollbar(self.get_image_address_in_folder(self.get_relative_path(), 0)[1])
        self.show_sub_frame(0)
        excel_file_address = self.get_relative_path(option="excel") + '/time-distance.xlsx'
        print(excel_file_address)
        x, y = self.open_excel_file(excel_file_address, 2)
        self.enable_graph()
        self.draw_graph(x, y, 2)

    def do_open_image_action(self):
        self.abs_path = QFileDialog.getExistingDirectory(None, "Open folder", "Outputs/")
        print("Path got folder:", self.abs_path)
        self.console_list.clear()
        self.disable_graph()
        self.mode = 1
        self.set_max_scrollbar(self.get_image_address_in_folder(self.get_relative_path(), 0)[1])
        image = QtGui.QPixmap(self.get_image_address_in_folder(self.get_relative_path(), 0)[0])
        self.main_frame.setPixmap(image.scaled(self.main_frame.size(), QtCore.Qt.KeepAspectRatio))
        self.show_sub_frame(0)

    def do_open_chart_action(self, direction):
        excel_file_address, _ = QFileDialog.getOpenFileName(None, "Open Excell File", 'OutPuts/', "(time-distance.xlsx)")
        print("Path got video:", self.abs_path)
        self.console_list.clear()
        self.mode = 1
        x, y = self.open_excel_file(excel_file_address, direction)
        self.enable_graph()
        self.clear_frame_box()
        self.draw_graph(x, y, direction)

    def open_excel_file(self, excel_file_address, direction):
        # To open Workbook
        wb = xlrd.open_workbook(excel_file_address)
        sheet = wb.sheet_by_index(0)
        a1_array = []
        a2_array = []
        try:
            info_list = []
            info_list.append(sheet.cell_value(1, 4))
            info_list.append(sheet.cell_value(2, 4))
            info_list.append(sheet.cell_value(3, 4))
            info_list.append(sheet.cell_value(4, 4))
            info_list.append(sheet.cell_value(5, 4))
            info_list.append(sheet.cell_value(6, 4))
            info_list.append(sheet.cell_value(7, 4))
            info_list.append(sheet.cell_value(8, 4))
            self.set_infomation_box(info_list)
        except:
            print("Have no information's patient")
            self.set_infomation_box(["", "", "", "Male", "", "", "", ""])
        # For row 0 and column 0
        if direction == 0: # X
            a1 = 0
            a2 = 1
        elif direction == 1: # Y
            a1 = 0
            a2 = 2
        elif direction == 2: # XY
            a1 = 1
            a2 = 2
        for i in range(1, sheet.nrows):
            a1_array.append(sheet.cell_value(i, a1))
            a2_array.append(sheet.cell_value(i, a2))
        return a1_array, a2_array

    def set_infomation_box(self, list):
        self.id_txt.setText(str(list[0]))
        self.name_txt.setText(str(list[1]))
        self.age_txt.setText(str(list[2]))
        self.sex_opt.setCurrentIndex(0 if list[3] == 'Male' else 1)
        self.height_txt.setText(str(list[4]))
        self.weight_txt.setText(str(list[5]))
        self.add_txt.setText(str(list[6]))
        self.note_txt.setText(str(list[7]))

    def get_relative_path(self, option=None):
        path_video = "Outputs/"
        is_path = False
        path = self.abs_path.split('/')
        for i in range(len(path) - 1):
            if path[i] == 'Outputs':
                is_path = True
                continue
            if is_path:
                path_video += path[i] + '/'
        if option == "excel":
            path_video += "excelFolder"
        else:
            path_video += path[-1]
        # print(path_video = "Outputs/exampole/1_1/imageFolder)
        return path_video

    def get_image_address_in_folder(self, path_folder, ind_image):
        img_address = []
        for filename in glob.glob(path_folder + '/*.jpg'):
            img_address.append(filename)
            # print(filename)
        return [img_address[ind_image], len(img_address)]


    def do_save_action(self):
        self.is_file_path()
        data = self.get_info_patient()
        main_process = MainProcess(file_name=self.abs_path)
        save_address = main_process.get_save_address()
        # workbook = xlrd.open_workbook(save_address)
        # worksheet = workbook.sheet_by_index(0)
        # current_row = worksheet.nrows
        # current_col = worksheet.ncols
        # print(save_address)
        wb = openpyxl.load_workbook(save_address)
        ws = wb['result']
        ws.cell(1, 4).value = "Patient's Infomation"
        for i in range(np.shape(data)[0]):
            ws.cell(i + 2, 4).value = str(data[i][0])
            ws.cell(i + 2, 5).value = str(data[i][1])
            # ws.write_row(4, i + 1, data[i])
        wb.save(save_address)
        wb.close()
        dlg = CustomDialog(type="completed", message="Lưu thông tin thành công!")
        if dlg.exec_():
            pass

    def show_finish_noti(self, status):
        if status:
            dlg = CustomDialog(type="completed", message="Quá trình xử lý hoàn tất!")
            if dlg.exec_():
                pass
            self.show_sub_frame(1)

    def update_console_log(self, list):
        self.console_list.addItem("{i} - {dis} cm".format(i=list[0], dis=list[1]))

    def do_calib_button(self, status):
        if status:
            main_process = MainProcess(file_name=self.abs_path)
            dlg = CustomDialog(type="completed", message="Quá trình xử lý hoàn tất!\r\nTrich xuat {number_image} khung hinh"
                               .format(number_image=main_process.get_index()))
            if dlg.exec_():
                pass
            self.console_list.addItem("Trich xuat {number_image} khung hinh".format(number_image=main_process.get_index()))

            image = main_process.get_frame(50)
            self.main_frame.show_main_frame(image)

            self.main_frame.is_laser = 0
            self.console_list.addItem("CHON 2 THONG SO MAU SAC TREN HINH")
            dlg = CustomDialog(
                message="CHON 2 THONG SO MAU SAC \r\n1. Chon thong so cho vung mau xanh \r\n2. Chon thong so cho vung laser")
            if not dlg.exec_():
                pass

    def show_sub_frame(self, ind_image):
        frame1, frame2, frame3, frame4 = self.get_sub_frame(ind_image)
        self.sub_frame1.setPixmap(frame1.scaled(self.sub_frame1.size(), QtCore.Qt.KeepAspectRatio))
        self.sub_frame2.setPixmap(frame2.scaled(self.sub_frame2.size(), QtCore.Qt.KeepAspectRatio))
        self.sub_frame3.setPixmap(frame3.scaled(self.sub_frame3.size(), QtCore.Qt.KeepAspectRatio))
        self.sub_frame4.setPixmap(frame4.scaled(self.sub_frame4.size(), QtCore.Qt.KeepAspectRatio))

    def clear_frame_box(self):
        self.sub_frame1.setText(" ")
        self.sub_frame2.setText(" ")
        self.sub_frame3.setText(" ")
        self.sub_frame4.setText(" ")
        self.main_frame.setText(" ")

    def show_subframe_to_mainframe(self, ind_image):
        self.disable_graph()
        image_address = ""
        if self.mode == 0:
            main_process = MainProcess(file_name=self.abs_path)
            image_address = main_process.get_image_address(ind_image)
        elif self.mode == 1:
            image_address = self.get_image_address_in_folder(self.get_relative_path(), ind_image)[0]

        image = QtGui.QPixmap(image_address)
        self.console_list.addItem(image_address)
        self.main_frame.setPixmap(image.scaled(self.main_frame.size(), QtCore.Qt.KeepAspectRatio))

    def get_sub_frame(self, start_p):
        self.sub_frame1.set_index_image(start_p)
        self.sub_frame2.set_index_image(start_p + 1)
        self.sub_frame3.set_index_image(start_p + 2)
        self.sub_frame4.set_index_image(start_p + 3)
        if self.mode == 0:
            main_process = MainProcess(file_name=self.abs_path)
            frame1 = QtGui.QPixmap(main_process.get_image_address(start_p))
            frame2 = QtGui.QPixmap(main_process.get_image_address(start_p + 1))
            frame3 = QtGui.QPixmap(main_process.get_image_address(start_p + 2))
            frame4 = QtGui.QPixmap(main_process.get_image_address(start_p + 3))
        elif self.mode == 1:
            frame1 = QtGui.QPixmap(self.get_image_address_in_folder(self.get_relative_path(), start_p)[0])
            frame2 = QtGui.QPixmap(self.get_image_address_in_folder(self.get_relative_path(), start_p + 1)[0])
            frame3 = QtGui.QPixmap(self.get_image_address_in_folder(self.get_relative_path(), start_p + 2)[0])
            frame4 = QtGui.QPixmap(self.get_image_address_in_folder(self.get_relative_path(), start_p + 3)[0])
        return frame1, frame2, frame3, frame4

    # def is_chart(self):
    #     if self.chart_rbtn.isChecked():
    #         self.console_list.addItem("Chart radio button is being checked")
    #         self.graphicsView.setGeometry(QtCore.QRect(216, 20, 531, 381))
    #         return True
    #     else:
    #         self.console_list.addItem("Chart radio button is NOT being checked")
    #         self.graphicsView.setGeometry(QtCore.QRect(0, 0, 0, 0))
    #         return False

    def enable_graph(self):
        # self.console_list.addItem("Graph is enabled")
        self.graphicsView.setGeometry(QtCore.QRect(216, 20, 531, 381))

    def disable_graph(self):
        # self.console_list.addItem("Graph is disabled")
        self.graphicsView.setGeometry(QtCore.QRect(0, 0, 0, 0))

    def draw_graph(self, x, y, direction=0):
        # x = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
        # y = [30, 32, 34, 32, 33, 31, 29, 32, 35, 45]
        x = np.array(x)
        y = np.array(y)
        if direction == 0 or direction == 1:
            x = x / 30
            y = y - y.mean()
        elif direction == 2:
            x = x - x.mean()
            y = y - y.mean()
        self.clear_graph()
        pen = pyqtgraph.mkPen(color=(255, 0, 0), width=0, style=QtCore.Qt.DashLine)
        self.graphicsView.plot(x, y, pen=pen, symbol='o', symbolSize=5, symbolBrush=('b'))

    def clear_graph(self):
        self.graphicsView.clear()

    def is_file_path(self):
        if os.path.exists(self.abs_path):
            pass
        else:
            dlg = CustomDialog(
                message="Chưa chọn tệp xử lý!", type='alert')
            if dlg.exec_():
                pass
            self.do_new_action()

    def calib_btn_clicked(self):
        if self.mode == 0:
            self.is_file_path()
            self.start_calib_thread()
            # print("Press CALIBRATE button")
            self.console_list.addItem("-----------------CALIBRATE----------------")
            self.console_list.addItem("Đang trích xuất khung hình")

    def run_btn_clicked(self):
        if self.mode == 0:
            self.is_file_path()
            self.start_run_thread()
            self.console_list.addItem("Blue color: " + str(self.main_frame.blue_HSV))
            self.console_list.addItem("Laser color: " + str(self.main_frame.laser_HSV))
            # print("Press RUN button")
            self.console_list.addItem("----------------------RUN button-----------------")
            self.enable_graph()


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
